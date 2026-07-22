from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import ssl
import subprocess
import sys
import urllib.request
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None

TMP_DEPS = Path("/private/tmp/acoustic_pydeps")
if TMP_DEPS.exists():
    sys.path.insert(0, str(TMP_DEPS))

ROOT = Path(__file__).resolve().parents[2]
RAW_ROOT = ROOT / "dataset" / "raw"
PROCESSED_ROOT = ROOT / "dataset" / "processed"


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def ensure_dirs(dataset_id: str) -> tuple[Path, Path]:
    raw_dir = RAW_ROOT / dataset_id
    processed_dir = PROCESSED_ROOT / dataset_id
    raw_dir.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(parents=True, exist_ok=True)
    return raw_dir, processed_dir


def download_file(url: str, dest: Path, expected_size: int | None = None, expected_sha256: str | None = None, insecure: bool = False) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists():
        size_ok = expected_size is None or dest.stat().st_size == expected_size
        sha_ok = expected_sha256 is None or sha256_file(dest) == expected_sha256
        if size_ok and sha_ok:
            print(f"skip existing: {dest.name}")
            return
        print(f"existing file does not match expected metadata, re-downloading: {dest}")
        dest.unlink()

    context = ssl._create_unverified_context() if insecure else None
    req = urllib.request.Request(url, headers={"User-Agent": "AcousticDataAudit/2026-06-29"})
    tmp = dest.with_suffix(dest.suffix + ".part")
    with urllib.request.urlopen(req, context=context, timeout=120) as resp, tmp.open("wb") as f:
        shutil.copyfileobj(resp, f, length=1024 * 1024)
    tmp.rename(dest)

    if expected_size is not None and dest.stat().st_size != expected_size:
        raise RuntimeError(f"size mismatch for {dest}: got {dest.stat().st_size}, expected {expected_size}")
    if expected_sha256 is not None:
        got = sha256_file(dest)
        if got != expected_sha256:
            raise RuntimeError(f"sha256 mismatch for {dest}: got {got}, expected {expected_sha256}")


def get_json(url: str):
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "AcousticDataAudit/2026-06-29",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.load(resp)


def write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        keys = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        fieldnames = keys
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_checksums(raw_dir: Path, processed_dir: Path) -> None:
    rows = []
    for p in sorted(raw_dir.rglob("*")):
        if p.is_file() and not p.name.endswith(".part"):
            rows.append(f"{sha256_file(p)}  {p.relative_to(raw_dir)}")
    (processed_dir / "checksums_sha256.txt").write_text("\n".join(rows) + ("\n" if rows else ""), encoding="utf-8")


def write_inventory(raw_dir: Path, processed_dir: Path) -> list[dict]:
    rows = []
    for p in sorted(raw_dir.rglob("*")):
        if p.is_file() and not p.name.endswith(".part"):
            rel = p.relative_to(raw_dir)
            rows.append({
                "relative_path": str(rel),
                "top_level": rel.parts[0] if rel.parts else "",
                "extension": p.suffix.lower(),
                "size_bytes": p.stat().st_size,
            })
    write_csv(processed_dir / "file_inventory.csv", rows, ["relative_path", "top_level", "extension", "size_bytes"])
    return rows


def extract_zip(zip_path: Path, target_dir: Path) -> None:
    marker = target_dir / ".extracted_complete"
    if marker.exists():
        print(f"skip extraction: {target_dir}")
        return
    target_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(target_dir)
    marker.write_text("ok\n", encoding="utf-8")


def extract_7z(first_part: Path, target_dir: Path) -> None:
    marker = target_dir / ".extracted_complete"
    if marker.exists():
        print(f"skip extraction: {target_dir}")
        return
    try:
        import multivolumefile
        import py7zr
    except Exception as exc:
        raise RuntimeError("py7zr and multivolumefile are required for HF_Lung_V1 multi-volume extraction") from exc
    target_dir.mkdir(parents=True, exist_ok=True)
    archive_base = first_part.with_suffix("")
    with multivolumefile.open(archive_base, mode="rb") as mv:
        with py7zr.SevenZipFile(mv, mode="r") as archive:
            archive.extractall(target_dir)
    marker.write_text("ok\n", encoding="utf-8")


def summarize_wavs(raw_dir: Path, processed_dir: Path) -> None:
    import wave

    groups = Counter()
    errors = []
    total = 0
    for wav in sorted(raw_dir.rglob("*.wav")):
        total += 1
        try:
            with wave.open(str(wav), "rb") as wf:
                key = (wf.getframerate(), wf.getnchannels(), wf.getsampwidth() * 8)
                groups[key] += 1
        except Exception as exc:
            errors.append({"relative_path": str(wav.relative_to(raw_dir)), "error": str(exc)})
    rows = [
        {
            "sample_rate_hz": sr,
            "channels": ch,
            "bit_depth": bd,
            "file_count": count,
            "total_wav_files": total,
            "confidence": "from wave headers" if count else "unknown / needs audit",
        }
        for (sr, ch, bd), count in sorted(groups.items())
    ]
    if not rows:
        rows = [{"sample_rate_hz": "unknown / needs audit", "channels": "unknown / needs audit", "bit_depth": "unknown / needs audit", "file_count": 0, "total_wav_files": total, "confidence": "no readable wav files found"}]
    write_csv(processed_dir / "audio_format_summary.csv", rows)
    if errors:
        write_csv(processed_dir / "audio_read_errors.csv", errors)


def write_markdown(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.strip() + "\n", encoding="utf-8")


def source_identity(dataset_id: str, processed_dir: Path, body: str) -> None:
    write_markdown(processed_dir / "source_identity.md", body)


def metadata_availability(processed_dir: Path, rows: list[dict]) -> None:
    write_csv(processed_dir / "metadata_availability.csv", rows, ["field", "availability", "evidence", "confidence"])


def split_risk(processed_dir: Path, text: str) -> None:
    write_markdown(processed_dir / "split_and_leakage_risk.md", text)


def mapping_readiness(processed_dir: Path, text: str) -> None:
    write_markdown(processed_dir / "mapping_readiness.md", text)


def audit_icbhi_2017() -> None:
    dataset_id = "icbhi_2017"
    raw_dir, processed_dir = ensure_dirs(dataset_id)
    files = [
        ("ICBHI_final_database.zip", "https://bhichallenge.med.auth.gr/sites/default/files/ICBHI_final_database/ICBHI_final_database.zip", 1978998275),
        ("events.zip", "https://bhichallenge.med.auth.gr/sites/default/files/ICBHI_final_database/events.zip", 479920),
        ("ICBHI_Challenge_demographic_information.txt", "https://bhichallenge.med.auth.gr/sites/default/files/ICBHI_final_database/ICBHI_Challenge_demographic_information.txt", 2586),
        ("ICBHI_Challenge_diagnosis.txt", "https://bhichallenge.med.auth.gr/sites/default/files/ICBHI_final_database/ICBHI_Challenge_diagnosis.txt", 1368),
        ("ICBHI_challenge_train_test.txt", "https://bhichallenge.med.auth.gr/sites/default/files/ICBHI_final_database/ICBHI_challenge_train_test.txt", 27219),
    ]
    for name, url, size in files:
        download_file(url, raw_dir / name, expected_size=size, insecure=True)
    source = raw_dir / "source_original"
    extract_zip(raw_dir / "ICBHI_final_database.zip", source / "ICBHI_final_database")
    extract_zip(raw_dir / "events.zip", source / "events")
    write_common_outputs(dataset_id, raw_dir, processed_dir)
    analyze_icbhi(raw_dir, processed_dir)


def audit_hf_lung_v1() -> None:
    dataset_id = "hf_lung_v1"
    raw_dir, processed_dir = ensure_dirs(dataset_id)
    parts = {
        **{f"train.7z.{i:03d}": 94371840 for i in range(1, 10)},
        "train.7z.010": 90182182,
        "test.7z.001": 94371840,
        "test.7z.002": 94371840,
        "test.7z.003": 46498282,
    }
    for name, size in parts.items():
        url = f"https://gitlab.com/techsupportHF/HF_Lung_V1/-/raw/master/{name}"
        download_file(url, raw_dir / name, expected_size=size)
    download_file("https://gitlab.com/techsupportHF/HF_Lung_V1/-/raw/master/README.md", raw_dir / "README.md")
    source = raw_dir / "source_original"
    extract_7z(raw_dir / "train.7z.001", source / "train")
    extract_7z(raw_dir / "test.7z.001", source / "test")
    write_common_outputs(dataset_id, raw_dir, processed_dir)
    analyze_hf_lung(raw_dir, processed_dir)


def audit_kauh_fraiwan() -> None:
    dataset_id = "kauh_fraiwan"
    raw_dir, processed_dir = ensure_dirs(dataset_id)
    meta_url = "https://data.mendeley.com/public-api/datasets/jwyy9np4gv/files?folder_id=root&version=3"
    files = get_json(meta_url)
    (processed_dir / "mendeley_file_metadata.json").write_text(json.dumps(files, indent=2), encoding="utf-8")
    for item in files:
        name = item["filename"]
        details = item["content_details"]
        download_file(details["download_url"], raw_dir / name, expected_size=details["size"], expected_sha256=details["sha256_hash"])
    source = raw_dir / "source_original"
    extract_zip(raw_dir / "Audio Files.zip", source / "audio_files")
    extract_zip(raw_dir / "Stethoscope Files.zip", source / "stethoscope_files")
    write_common_outputs(dataset_id, raw_dir, processed_dir)
    analyze_kauh(raw_dir, processed_dir)


def audit_sprsound() -> None:
    dataset_id = "sprsound"
    raw_dir, processed_dir = ensure_dirs(dataset_id)
    commit = "874eeb8736ddb78937c2fb5332fc7e7293d0f0ca"
    url = f"https://codeload.github.com/SJTU-YONGFU-RESEARCH-GRP/SPRSound/zip/{commit}"
    archive = raw_dir / f"SPRSound-{commit}.zip"
    download_file(url, archive)
    source = raw_dir / "source_original"
    extract_zip(archive, source)
    write_common_outputs(dataset_id, raw_dir, processed_dir)
    analyze_sprsound(raw_dir, processed_dir, commit)


def write_common_outputs(dataset_id: str, raw_dir: Path, processed_dir: Path) -> None:
    write_checksums(raw_dir, processed_dir)
    write_inventory(raw_dir, processed_dir)
    summarize_wavs(raw_dir, processed_dir)


def analyze_icbhi(raw_dir: Path, processed_dir: Path) -> None:
    ann_files = [p for p in (raw_dir / "source_original").rglob("*.txt") if re.match(r"^\d+_", p.name)]
    dist = Counter()
    for p in ann_files:
        for line in p.read_text(errors="ignore").splitlines():
            parts = line.split()
            if len(parts) >= 4:
                crackle, wheeze = parts[2], parts[3]
                label = "both" if crackle == "1" and wheeze == "1" else "crackle" if crackle == "1" else "wheeze" if wheeze == "1" else "normal"
                dist[label] += 1
    write_csv(processed_dir / "label_schema.csv", [
        {"raw_label": "crackle_flag,wheeze_flag", "label_unit": "respiratory_cycle", "meaning": "0/1 flags in per-recording annotation txt", "source": "ICBHI official files"}
    ])
    write_csv(processed_dir / "class_distribution.csv", [{"label": k, "count": v, "unit": "respiratory_cycle"} for k, v in sorted(dist.items())])
    source_identity("icbhi_2017", processed_dir, """
# ICBHI 2017 Source Identity

Official source: https://bhichallenge.med.auth.gr/ICBHI_2017_Challenge

Access/license: official challenge page says the database is freely available for research. No separate license file was found in this first-pass download.

Downloaded files: official main zip plus `events.zip`, demographics, diagnosis, and train/test split side files. Source extraction is under `dataset/raw/icbhi_2017/source_original/`.

Confidence: high for source identity and official file route.
""")
    metadata_availability(processed_dir, [
        {"field": "patient_id", "availability": "available", "evidence": "filename first element and diagnosis/demographic files", "confidence": "high"},
        {"field": "age", "availability": "available", "evidence": "ICBHI_Challenge_demographic_information.txt", "confidence": "high"},
        {"field": "sex", "availability": "available", "evidence": "ICBHI_Challenge_demographic_information.txt", "confidence": "high"},
        {"field": "disease", "availability": "available", "evidence": "ICBHI_Challenge_diagnosis.txt", "confidence": "high"},
        {"field": "recording_location", "availability": "available", "evidence": "filename chest-location code", "confidence": "high"},
        {"field": "device", "availability": "available", "evidence": "filename recording-equipment code", "confidence": "high"},
        {"field": "session_id", "availability": "unknown / needs audit", "evidence": "filename has recording index but not explicit session semantics", "confidence": "medium"},
        {"field": "official_split", "availability": "available", "evidence": "ICBHI_challenge_train_test.txt", "confidence": "high"},
    ])
    split_risk(processed_dir, """
# ICBHI Split And Leakage Risk

Official train/test split is available. Patient ID, chest location, acquisition mode, and device are encoded in filenames. Leakage audit must confirm patient independence and device/site imbalance before using official split as the only benchmark split.
""")
    mapping_readiness(processed_dir, """
# ICBHI Mapping Readiness

- Normal/abnormal: yes. Cycles with no crackle and no wheeze map to normal; any crackle or wheeze maps to abnormal.
- Crackle/wheeze/both/normal: yes. The raw label flags directly support this.
- Dataset-specific head: yes, four-class cycle head plus optional disease-level subject head if diagnosis labels are explicitly used.
- Disease/event separation: separate. Event labels are cycle-level crackle/wheeze flags; disease labels are subject-level diagnosis file entries.
""")


def analyze_hf_lung(raw_dir: Path, processed_dir: Path) -> None:
    label_files = sorted((raw_dir / "source_original").rglob("*_label.txt"))
    token_counts = Counter()
    for p in label_files:
        for line in p.read_text(errors="ignore").splitlines():
            parts = line.strip().split()
            if parts:
                token_counts[parts[0]] += 1
    known = [
        ("inhalation", "event", "breath phase label reported by README/PLOS paper"),
        ("exhalation", "event", "breath phase label reported by README/PLOS paper"),
        ("continuous_adventitious_sound", "event", "CAS includes wheeze, stridor, rhonchi"),
        ("discontinuous_adventitious_sound", "event", "DAS labels are crackles"),
    ]
    write_csv(processed_dir / "label_schema.csv", [{"raw_label": k, "label_unit": u, "meaning": m, "source": "HF_Lung_V1 README/PLOS paper"} for k, u, m in known])
    rows = [{"label": k, "count": v, "unit": "label_interval"} for k, v in token_counts.most_common()]
    if not rows:
        rows = [{"label": "unknown / needs audit", "count": 0, "unit": "label_files_not_found_or_unparsed"}]
    write_csv(processed_dir / "class_distribution.csv", rows)
    source_identity("hf_lung_v1", processed_dir, """
# HF_Lung_V1 Source Identity

Official source: https://gitlab.com/techsupportHF/HF_Lung_V1

Access/license: public GitLab repository. README states CC BY 4.0.

Downloaded files: 13 split 7z archives plus README. Source extraction is under `dataset/raw/hf_lung_v1/source_original/`.

Confidence: high for source identity and downloaded route; medium for parsed class counts because label TXT semantics require a second review against the paper/code.
""")
    metadata_availability(processed_dir, [
        {"field": "patient_id", "availability": "unknown / needs audit", "evidence": "README says same-date files likely same subject but no direct patient ID field observed in first pass", "confidence": "medium"},
        {"field": "age", "availability": "unknown / needs audit", "evidence": "not found in first-pass file/README inspection", "confidence": "medium"},
        {"field": "sex", "availability": "unknown / needs audit", "evidence": "not found in first-pass file/README inspection", "confidence": "medium"},
        {"field": "disease", "availability": "unknown / needs audit", "evidence": "not found in first-pass file/README inspection", "confidence": "medium"},
        {"field": "recording_location", "availability": "partially available", "evidence": "HF-Type-1 filenames include L1-L8 location; steth_ files should not infer location per README", "confidence": "high"},
        {"field": "device", "availability": "available", "evidence": "filename prefixes steth_ and trunc_; README describes Littmann 3200 and HF-Type-1", "confidence": "high"},
        {"field": "session_id", "availability": "partially available", "evidence": "date/time encoded in filename; README warns same date likely same subject", "confidence": "high"},
        {"field": "official_split", "availability": "available", "evidence": "train/test folders from study release", "confidence": "high"},
    ])
    split_risk(processed_dir, """
# HF_Lung_V1 Split And Leakage Risk

The source ships train/test folders, but README warns that files with the same date are very likely from the same subject and should not be assigned to different train/validation/test sets. Any derived split must group by date/session and device prefix.
""")
    mapping_readiness(processed_dir, """
# HF_Lung_V1 Mapping Readiness

- Normal/abnormal: likely, but needs interval-level definition. CAS/DAS intervals can indicate abnormal adventitious events; unlabeled background is not automatically normal without protocol review.
- Crackle/wheeze/both/normal: partial/lossy. Crackles are DAS; wheeze is one CAS subtype, but CAS also includes stridor/rhonchi.
- Dataset-specific head: yes, event-detection heads for inhalation/exhalation/CAS/DAS or subtype heads.
- Disease/event separation: mostly event labels; disease labels were not found in first pass.
""")


def analyze_kauh(raw_dir: Path, processed_dir: Path) -> None:
    xlsx = raw_dir / "Data annotation.xlsx"
    rows_out = []
    dist_rows = []
    if openpyxl and xlsx.exists():
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(v).strip() if v is not None else f"col_{i}" for i, v in enumerate(rows[0])]
        for idx, h in enumerate(headers):
            vals = [r[idx] for r in rows[1:] if idx < len(r) and r[idx] not in (None, "")]
            unique = Counter(str(v).strip() for v in vals)
            rows_out.append({"column": h, "non_missing": len(vals), "unique_values": len(unique), "sample_values": "; ".join(list(unique.keys())[:10])})
            if any(key in h.lower() for key in ["sound", "diagn", "class", "label"]):
                for label, count in unique.items():
                    dist_rows.append({"label": label, "count": count, "unit": f"xlsx_column:{h}"})
    write_csv(processed_dir / "label_schema.csv", rows_out or [{"column": "unknown / needs audit", "non_missing": 0, "unique_values": 0, "sample_values": ""}])
    write_csv(processed_dir / "class_distribution.csv", dist_rows or [{"label": "unknown / needs audit", "count": 0, "unit": "xlsx_not_parsed"}])
    source_identity("kauh_fraiwan", processed_dir, """
# KAUH / Fraiwan Source Identity

Official source: https://data.mendeley.com/datasets/jwyy9np4gv/3

Access/license: Mendeley Data version 3, CC BY 4.0. The dataset page title is "A dataset of lung sounds recorded from the chest wall using an electronic stethoscope".

Downloaded files: `Audio Files.zip`, `Data annotation.xlsx`, and `Stethoscope Files.zip`. Source extraction is under `dataset/raw/kauh_fraiwan/source_original/`.

Confidence: high for Mendeley source identity and file hashes. Project-level identity as KAUH/Fraiwan source family should still be confirmed by Project Management/Jingping.
""")
    metadata_availability(processed_dir, [
        {"field": "patient_id", "availability": "available", "evidence": "Mendeley description says filenames encode patient number P1-P112", "confidence": "high"},
        {"field": "age", "availability": "unknown / needs audit", "evidence": "not identified in first-pass xlsx summary", "confidence": "medium"},
        {"field": "sex", "availability": "unknown / needs audit", "evidence": "not identified in first-pass xlsx summary", "confidence": "medium"},
        {"field": "disease", "availability": "available", "evidence": "Mendeley description and annotation workbook include diagnosis semantics", "confidence": "high"},
        {"field": "recording_location", "availability": "available", "evidence": "Mendeley description says annotation includes chest-wall location", "confidence": "high"},
        {"field": "device", "availability": "available", "evidence": "3M Littmann Electronic Stethoscope model 3200", "confidence": "high"},
        {"field": "session_id", "availability": "unknown / needs audit", "evidence": "not identified in first-pass source description", "confidence": "medium"},
        {"field": "official_split", "availability": "not_provided", "evidence": "no official split found in first-pass source files", "confidence": "medium"},
    ])
    split_risk(processed_dir, """
# KAUH / Fraiwan Split And Leakage Risk

No official split was found in the first-pass source files. Any benchmark split should group by patient number and stratify by diagnosis/sound type and chest location where possible.
""")
    mapping_readiness(processed_dir, """
# KAUH / Fraiwan Mapping Readiness

- Normal/abnormal: likely, using sound-type labels after annotation workbook review.
- Crackle/wheeze/both/normal: partial. Wheeze/crackle/normal sound types are present, but both-label semantics require audit.
- Dataset-specific head: yes, sound-type head and possibly disease diagnosis head.
- Disease/event separation: mixed in the annotation workbook/source description but separable by column; do not conflate diagnosis with sound event labels.
""")


def analyze_sprsound(raw_dir: Path, processed_dir: Path, commit: str) -> None:
    json_files = sorted((raw_dir / "source_original").rglob("*.json"))
    record_counts = Counter()
    event_counts = Counter()
    for p in json_files:
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        rec = data.get("record_annotation")
        if rec:
            record_counts[str(rec)] += 1
        for ev in data.get("event_annotation") or []:
            t = ev.get("type")
            if t:
                event_counts[str(t)] += 1
    label_rows = [
        {"raw_label": "Normal/CAS/DAS/CAS & DAS/Poor Quality", "label_unit": "record", "meaning": "record-level annotation", "source": "SPRSound README/JSON"},
        {"raw_label": "Normal/Rhonchi/Wheeze/Stridor/Coarse Crackle/Fine Crackle/Wheeze+Crackle", "label_unit": "event", "meaning": "event-level annotation", "source": "SPRSound README/JSON"},
    ]
    write_csv(processed_dir / "label_schema.csv", label_rows)
    dist = [{"label": k, "count": v, "unit": "record_annotation"} for k, v in sorted(record_counts.items())]
    dist += [{"label": k, "count": v, "unit": "event_annotation"} for k, v in sorted(event_counts.items())]
    write_csv(processed_dir / "class_distribution.csv", dist or [{"label": "unknown / needs audit", "count": 0, "unit": "json_not_parsed"}])
    source_identity("sprsound", processed_dir, f"""
# SPRSound Source Identity

Official source: https://github.com/SJTU-YONGFU-RESEARCH-GRP/SPRSound

Access/license: public GitHub repository with CC BY 4.0 license metadata. Downloaded full repository zip pinned to commit `{commit}`.

Downloaded file: `SPRSound-{commit}.zip`. Source extraction is under `dataset/raw/sprsound/source_original/`.

Confidence: high for GitHub source and commit pin. Paper/DataPort version alignment remains a separate citation audit item.
""")
    metadata_availability(processed_dir, [
        {"field": "patient_id", "availability": "available", "evidence": "filename first element", "confidence": "high"},
        {"field": "age", "availability": "available", "evidence": "filename second element", "confidence": "high"},
        {"field": "sex", "availability": "available", "evidence": "filename gender element", "confidence": "high"},
        {"field": "disease", "availability": "unknown / needs audit", "evidence": "patient summaries may include additional fields but disease labels not verified in first pass", "confidence": "medium"},
        {"field": "recording_location", "availability": "available", "evidence": "filename location element p1-p8", "confidence": "high"},
        {"field": "device", "availability": "available", "evidence": "README states Yunting model II Stethoscopes", "confidence": "high"},
        {"field": "session_id", "availability": "available", "evidence": "filename recording number", "confidence": "medium"},
        {"field": "official_split", "availability": "available", "evidence": "challenge-year directories and inter/intra test JSON folders", "confidence": "high"},
    ])
    split_risk(processed_dir, """
# SPRSound Split And Leakage Risk

The repository contains multiple challenge-year/version directories. BioCAS2022 includes train/test and inter/intra test subsets; later challenge directories add additional test or detection/compression material. A phase-1 benchmark must pin which directories are included and group by patient number to avoid repeated-patient leakage.
""")
    mapping_readiness(processed_dir, """
# SPRSound Mapping Readiness

- Normal/abnormal: yes for event labels and likely for record labels, with Poor Quality treated separately.
- Crackle/wheeze/both/normal: partial. Event labels distinguish wheeze, fine/coarse crackle, and Wheeze+Crackle; rhonchi/stridor are other adventitious sounds and require an explicit mapping policy.
- Dataset-specific head: yes, event-level and record-level heads.
- Disease/event separation: event labels are respiratory sound classes; disease labels were not verified in first pass.
""")


def build_phase1_summary() -> None:
    readiness = {
        "icbhi_2017": {
            "normal_abnormal": "yes",
            "crackle_wheeze_both_normal": "yes",
            "dataset_specific_head": "yes",
            "leakage_warning": "confirm official split patient independence and device/site balance",
        },
        "hf_lung_v1": {
            "normal_abnormal": "likely / needs interval policy",
            "crackle_wheeze_both_normal": "partial / lossy",
            "dataset_specific_head": "yes",
            "leakage_warning": "group by same recording date/session and device prefix",
        },
        "kauh_fraiwan": {
            "normal_abnormal": "likely / needs workbook audit",
            "crackle_wheeze_both_normal": "partial / needs both-label semantics",
            "dataset_specific_head": "yes",
            "leakage_warning": "no official split found; group by patient number",
        },
        "sprsound": {
            "normal_abnormal": "yes, with Poor Quality separate",
            "crackle_wheeze_both_normal": "partial / explicit rhonchi-stridor policy needed",
            "dataset_specific_head": "yes",
            "leakage_warning": "pin included challenge directories and group by patient",
        },
    }

    rows = []
    for dataset_id in ["icbhi_2017", "hf_lung_v1", "kauh_fraiwan", "sprsound"]:
        processed = PROCESSED_ROOT / dataset_id
        inv = processed / "file_inventory.csv"
        raw = RAW_ROOT / dataset_id
        file_count = 0
        if inv.exists():
            with inv.open(newline="", encoding="utf-8") as f:
                file_count = sum(1 for _ in csv.DictReader(f))
        audio_rows = []
        audio_path = processed / "audio_format_summary.csv"
        if audio_path.exists():
            with audio_path.open(newline="", encoding="utf-8") as f:
                audio_rows = list(csv.DictReader(f))
        wav_count = audio_rows[0].get("total_wav_files", "unknown / needs audit") if audio_rows else "unknown / needs audit"
        audio_formats = "; ".join(
            f"{r.get('sample_rate_hz')} Hz/{r.get('channels')} ch/{r.get('bit_depth')} bit: {r.get('file_count')}"
            for r in audio_rows
        ) or "unknown / needs audit"
        rdy = readiness[dataset_id]
        rows.append({
            "dataset_id": dataset_id,
            "raw_dir": str(raw.relative_to(ROOT)),
            "processed_dir": str(processed.relative_to(ROOT)),
            "file_count": file_count,
            "wav_count": wav_count,
            "audio_formats": audio_formats,
            "normal_abnormal_mapping": rdy["normal_abnormal"],
            "crackle_wheeze_both_normal_mapping": rdy["crackle_wheeze_both_normal"],
            "dataset_specific_head": rdy["dataset_specific_head"],
            "leakage_warning": rdy["leakage_warning"],
            "checksums": str((processed / "checksums_sha256.txt").relative_to(ROOT)),
            "class_distribution": str((processed / "class_distribution.csv").relative_to(ROOT)),
            "mapping_readiness": str((processed / "mapping_readiness.md").relative_to(ROOT)),
        })
    write_csv(PROCESSED_ROOT / "phase1_dataset_matrix_2026-06-29.csv", rows)
    md = ["# Phase-1 Dataset Audit Summary - 2026-06-29", "", "Downloaded/audited datasets: " + ", ".join(r["dataset_id"] for r in rows), ""]
    md.extend([
        "## Cross-Dataset Verdict",
        "",
        "- All four phase-1 datasets were downloaded through the approved source routes and audited into the simplified `dataset/raw`, `dataset/processed`, and `dataset/script` workspace.",
        "- ICBHI and SPRSound can directly support respiratory-event style heads; HF_Lung and KAUH/Fraiwan need explicit mapping policy before a shared four-class crackle/wheeze/both/normal benchmark.",
        "- Patient/session grouping is the main cross-dataset leakage risk. HF_Lung same-date files, KAUH patient IDs, SPRSound patient IDs, and ICBHI patient IDs must be treated as split-group keys.",
        "- Do not build final model-ready manifests until label mapping and split grouping policies are approved.",
        "",
    ])
    for r in rows:
        md.append(f"## {r['dataset_id']}")
        md.append(f"- Raw dir: `{r['raw_dir']}`")
        md.append(f"- Processed dir: `{r['processed_dir']}`")
        md.append(f"- Inventory file count: {r['file_count']}")
        md.append(f"- WAV count: {r['wav_count']}")
        md.append(f"- Audio formats: {r['audio_formats']}")
        md.append(f"- Normal/abnormal mapping: {r['normal_abnormal_mapping']}")
        md.append(f"- Crackle/wheeze/both/normal mapping: {r['crackle_wheeze_both_normal_mapping']}")
        md.append(f"- Leakage warning: {r['leakage_warning']}")
        md.append(f"- Class distribution: `{r['class_distribution']}`")
        md.append(f"- Mapping readiness: `{r['mapping_readiness']}`")
        md.append("")
    (PROCESSED_ROOT / "phase1_dataset_audit_summary_2026-06-29.md").write_text("\n".join(md), encoding="utf-8")


def main(argv: list[str]) -> None:
    if len(argv) != 2:
        raise SystemExit("usage: phase1_dataset_workflow.py [icbhi_2017|hf_lung_v1|kauh_fraiwan|sprsound|summary|all]")
    target = argv[1]
    funcs = {
        "icbhi_2017": audit_icbhi_2017,
        "hf_lung_v1": audit_hf_lung_v1,
        "kauh_fraiwan": audit_kauh_fraiwan,
        "sprsound": audit_sprsound,
        "summary": build_phase1_summary,
    }
    if target == "all":
        for name in ["kauh_fraiwan", "hf_lung_v1", "icbhi_2017", "sprsound"]:
            print(f"=== {name} ===")
            funcs[name]()
        build_phase1_summary()
    else:
        funcs[target]()


if __name__ == "__main__":
    main(sys.argv)
